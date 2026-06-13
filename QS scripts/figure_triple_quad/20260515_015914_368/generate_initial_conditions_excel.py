# -*- coding: utf-8 -*-
"""Generate Excel table documenting initial condition settings for numerical experiments."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 三列：状态变量/参数、默认值、单位
ROWS = [
    (
        "扫描统一性（所有 ODE 组内）",
        "各扫描组内固定不变",
        "—",
    ),
    (
        "菌体密度 N1(0), N2(0), N3(0)",
        "0.12（各菌相同）",
        "无量纲（OD 丰度）",
    ),
    (
        "营养底物 S(0)",
        "1.0（= S0）",
        "无量纲（S/S0）",
    ),
    (
        "信号 A1(0), A2(0), A3(0)",
        "0.02（各信号相同）",
        "无量纲浓度",
    ),
    (
        "产物 B1(0), B2(0), B3(0)",
        "1×10⁻⁶（各产物相同）",
        "无量纲浓度",
    ),
    (
        "累积资源 R(0)",
        "0",
        "无量纲（累积积分量）",
    ),
    (
        "迟滞门控 q1(0), q2(0), q3(0)（迟滞实验）",
        "0（各门控相同）",
        "无量纲（0–1）",
    ),
    (
        "四档主实验（hyst/nohyst × Binhibit/Bpromote）",
        "上述初值均不变",
        "同上行",
    ),
]


def _style_header(ws, ncols: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["状态变量/参数", "默认值", "单位"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 28


def _auto_width(ws, ncols: int, nrows: int) -> None:
    widths = [28, 22, 22]
    for i, w in enumerate(widths[:ncols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)
    for r in range(2, nrows + 1):
        ws.row_dimensions[r].height = 60
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = wrap
            cell.border = border


def main() -> None:
    out = Path(__file__).resolve().parent / "数值实验初始条件设置依据.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "初始条件设置依据"

    ncols = 3
    _style_header(ws, ncols)

    for idx, row in enumerate(ROWS, 1):
        for col, val in enumerate(row, 1):
            ws.cell(row=idx + 1, column=col, value=val)

    _auto_width(ws, ncols, len(ROWS) + 1)
    ws.freeze_panes = "A2"

    wb.save(out)
    print(f"saved: {out}")


if __name__ == "__main__":
    main()
